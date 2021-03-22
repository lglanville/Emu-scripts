from datetime import datetime
import calendar
import re
import argparse
from pathlib import Path


MONTHS = list(calendar.month_name)[1:]
MONTHS.extend(list(calendar.month_abbr)[1:])
MONTHGROUP = "(?P<MONTH>" + '|'.join([m for m in MONTHS]) + r")"
MONTHGROUP2 = "(?:(?P<MONTH2>" + '|'.join([m for m in MONTHS]) + r")[ ,./-]{1,3})"
MONTH_NUMS = [str(m) for m in list(range(1, 13, 1))]
MONTHNUM_GROUP = "(?P<MONTHNUM>" + '|'.join([m for m in MONTH_NUMS]) + ")"
DAY = r"(?P<DAY>[0-3]?\d)"
DAY2 = r"(?:(?P<DAY2>[0-3]?\d)[ ,./-]{1,3})"
DELIM = "[ ,./]{1,3}"
RDELIM = "[ ,./-]{1,3}"
YEAR = r"(?P<YEAR>(15|16|17|18|19|20)\d{2})"
CIRCA = r"(?P<CIRCA>c\.? ?|circa ?)?"

DAY_RANGE_RE = re.compile('('+DAY2+'?'+DAY+DELIM+MONTHGROUP+DELIM+YEAR+')', flags=re.IGNORECASE)
REV_DAY_RANGE_RE = re.compile('('+MONTHGROUP+DELIM+DAY2+'?'+DAY+DELIM+YEAR+')', flags=re.IGNORECASE)
MONTH_RANGE_RE = re.compile('('+MONTHGROUP2+'?'+MONTHGROUP+DELIM+YEAR+')', flags=re.IGNORECASE)
DELIM_RE = re.compile('('+DAY+DELIM+MONTHNUM_GROUP+DELIM+YEAR+')', flags=re.IGNORECASE)
MDELIM_RE = re.compile('('+CIRCA+MONTHNUM_GROUP+DELIM+YEAR+')', flags=re.IGNORECASE)
YEAR_RE = re.compile(r"("+CIRCA+YEAR+"(?P<DECADE>s)?)", flags=re.IGNORECASE)

DAYS = calendar.day_name[:]
DAYS.extend(calendar.day_abbr[:])
EARLIEST = 1800
LATEST = datetime.now().year

class circadate(datetime):
    """Class to express fudgy dates"""

    def __new__(cls, year, month=1, day=1, hour=0, minute=0, second=0, circa=False, no_day=False, no_month=False):
        if LATEST > year > EARLIEST:
            new = super().__new__(cls, year, month, day)
            new.circa = circa
            new.no_day = no_day
            new.no_month = no_month
            return new

    def __str__(self):
        if self.no_day:
            ds = self.strftime("%B %Y")
        elif self.no_month:
            ds = self.strftime("%Y")
        else:
            ds = self.strftime("%d %B %Y").lstrip('0')
        if self.circa:
            ds = 'c.'+ds
        return ds

    def __repr__(self):
        return f"<circadate {str(self)}>"

    @classmethod
    def strptime(cls, date_string, format, circa=False, no_day=False, no_month=False):
        import _strptime
        try:
            new = _strptime._strptime_datetime(cls, date_string, format)
            new.circa = circa
            new.no_day = no_day
            new.no_month = no_month
            if LATEST > new.year > EARLIEST:
                return new
        except Exception:
            print("Failed to convert", date_string, "to date")
            return None


class daterange(object):
    """Class to express a date range"""

    def __init__(self, *dates):
        self.earliest = min(dates)
        self.latest = max(dates)

    def __str__(self):
        if str(self.earliest) == str(self.latest):
            return str(self.earliest)
        else:
            if str(self.earliest) in str(self.latest):
                return str(self.latest)
            elif str(self.latest) in str(self.earliest):
                return str(self.earliest)
            else:
                return str(self.earliest)+'-'+str(self.latest)

    def latestcolumn(self, circa_margin=5):
        if self.latest.circa and self.latest.no_month:
            return self.latest.year+circa_margin
        else:
            return self.latest.year

    def earliestcolumn(self, circa_margin=5):
        if self.earliest.circa and self.earliest.no_month:
            return self.earliest.year-circa_margin
        else:
            return self.earliest.year

    def columns(self, circa_margin=5):
        return (
            str(self), self.earliestcolumn(circa_margin),
            self.latestcolumn(circa_margin))


def strip_days(text):
    for day in DAYS:
        text = text.replace(day, '')
    while '  ' in text:
        text = text.replace('  ', ' ')
    text = text.replace(b'\xe2\x80\x93'.decode(), '-')
    return text


def pull_dates(text, circa=False):
    """Extract all potential dates from a block of text and return a
    daterange object.
    """
    dates = []
    d, text = dayrange_extract(text)
    dates.extend(d)
    d, text = reverse_dayrange_extract(text)
    dates.extend(d)
    d, text = monthrange_extract(text)
    dates.extend(d)
    d, text = delimited_extract(text)
    dates.extend(d)
    d, text = monthdelimited_extract(text)
    dates.extend(d)
    d, text = year_extract(text)
    dates.extend(d)
    dates = [d for d in dates if d is not None]
    if dates != []:
        return daterange(*dates)

def norm_day(day):
    return '0' * (2-len(day))+day


def dayrange_extract(text):
    """Extracts dates from text in the form of  12 January 1982, 1-12 Jan,1982
    etc"""
    dates = []
    for date in DAY_RANGE_RE.finditer(text):
        vals = date.groupdict()
        fstring = "%d %b %Y"
        dstring = f"{norm_day(vals['DAY'])} {vals['MONTH'][:3].title()} {vals['YEAR']}"
        dates.append(circadate.strptime(dstring, fstring))
        if vals['DAY2'] is not None:
            dstring = f"{norm_day(vals['DAY2'])} {vals['MONTH'][:3].title()} {vals['YEAR']}"
            dates.append(circadate.strptime(dstring, fstring))
            text = text.replace(date[0], '')
    return dates, text


def reverse_dayrange_extract(text):
    """Extracts dates in  the form of jan 6 1978, aug 8-9 1876 etc."""
    dates = []
    for date in REV_DAY_RANGE_RE.finditer(text):
        vals = date.groupdict()
        fstring = "%d %b %Y"
        dstring = f"{norm_day(vals['DAY'])} {vals['MONTH'][:3].title()} {vals['YEAR']}"
        dates.append(circadate.strptime(dstring, fstring))
        if vals['DAY2'] is not None:
            dstring = f"{norm_day(vals['DAY2'])} {vals['MONTH'][:3].title()} {vals['YEAR']}"
            dates.append(circadate.strptime(dstring, fstring))
            text = text.replace(date[0], '')
    return dates, text


def monthrange_extract(text):
    """Extracts dates from text in the form of  January 1982, Jan, 1982,
    aug-sep 1765 etc"""
    dates = []
    for date in MONTH_RANGE_RE.finditer(text):
        vals = date.groupdict()
        fstring = "%b %Y"
        dstring = f"{vals['MONTH'][:3].title()} {vals['YEAR']}"
        dates.append(circadate.strptime(dstring, fstring, no_day=True))
        if vals['MONTH2'] is not None:
            dstring = f"{vals['MONTH2'][:3].title()} {vals['YEAR']}"
            dates.append(circadate.strptime(dstring, fstring, no_day=True))
        text = text.replace(date[0], '')
    return dates, text


def delimited_extract(text):
    """Extracts dates from text in the form of  1/1/1982, 1.1.1982 etc"""
    dates = []
    for date in DELIM_RE.finditer(text):
        vals = date.groupdict()
        dates.append(
            circadate(
                int(vals['YEAR']),
                int(vals['MONTHNUM']),
                int(vals['DAY'])))
        text = text.replace(date[0], '')
    return dates, text


def monthdelimited_extract(text):
    """Extracts dates from text in the form of  12/1982, 6.1982 etc"""
    dates = []
    for date in MDELIM_RE.finditer(text):
        vals = date.groupdict()
        dates.append(circadate(int(vals['YEAR']), int(vals['MONTHNUM']), no_day=True))
        text = text.replace(date[0], '')
    return dates, text


def year_extract(text):
    """Extracts a year or decade from text in the form of 1982, c.1982, 1980s
    etc"""
    dates = []
    for date in YEAR_RE.finditer(text):
        vals = date.groupdict()
        circa = False
        if vals['CIRCA'] is not None:
            circa = True
        if vals['DECADE'] is not None:
            earliest = int(vals['YEAR'])
            latest = earliest + 10
            dates.append(circadate(earliest, no_month=True))
            dates.append(circadate(latest, no_month=True))
        else:
            dates.append(circadate(int(vals['YEAR']), no_month=True, circa=circa))
    return dates, text


def main(workbookpath, column='EADUnitDate', cmargin=5):
    try:
        import openpyxl
    except ImportError:
        print("You don't have openpyxl installed. Try 'pip install openpyxl'")
        exit()
    wb = openpyxl.open(workbookpath)
    for ws in wb.worksheets:
        date_column = None
        for cell in ws[1]:
            if cell.value == column:
                date_column = cell.column_letter
                col_index = cell.column
                break
        if date_column is not None:
            ws.insert_cols(col_index+1, 3)
            ws.cell(1, col_index+1).value = 'EADUnitDate'
            ws.cell(1, col_index+2).value = 'EADUnitDateEarliest'
            ws.cell(1, col_index+3).value = 'EADUnitDateLatest'
            for cell in ws[date_column][1:]:
                if cell.value is not None:
                    if type(cell.value) == datetime:
                        date = cell.value
                        dates = daterange(circadate(date.year, date.month, date.day))
                    else:
                        datestring = str(cell.value).replace('\n', ' ')
                        datestring = strip_days(datestring)
                        dates = pull_dates(datestring)
                    if dates is not None:
                        cols = dates.columns(circa_margin=cmargin)
                        ds = str(cell.value).replace('\n', ' ')
                        print(f"Row {cell.row}: Converted", ds, "->", *cols)
                        ws.cell(cell.row, col_index+1).value = cols[0]
                        ws.cell(cell.row, col_index+2).value = cols[1]
                        ws.cell(cell.row, col_index+3).value = cols[2]
                    else:
                        print(f"Row {cell.row}: Found no dates in ", datestring)
                        ws.cell(cell.row, col_index+1).value = 'Undated'
                else:
                    print(f"Row {cell.row}: Empty cell")
                    ws.cell(cell.row, col_index+1).value = 'Undated'

    p = Path(workbookpath)
    new_path = p.parent / (p.stem+'_datefixed.xlsx')
    print("Saving workbook to", new_path)
    wb.save(new_path)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Fix horrible dates in an Excel spreadsheet')
    parser.add_argument(
        'workbook', metavar='i', type=str, help='Workbook with horrible dates')
    parser.add_argument(
        '--datecolumn', '-d', default='EADUnitDate',
        help='Column to extract dates from, defaults to EADUnitDate')
    parser.add_argument(
        '--circamargin', '-c', type=int, default=5,
        help='Margin to place on earliest and latest dates where dates are circa')
    parser.add_argument(
        '--earliest', '-e', type=int,
        help='Earliest possible year for dates extracted')
    args = parser.parse_args()
    if args.earliest is not None:
        EARLIEST = args.earliest
    main(args.workbook, column=args.datecolumn, cmargin=args.circamargin)
